"""XLSX generator. Builds Excel workbooks from typed Sheet data.

SOLID: XlsxStyleConfig for injectable styles (OCP), Sheet dataclass for typed
data (SRP). Tabular domain -- single output format, no Renderer Protocol needed.
"""

from __future__ import annotations

from dataclasses import dataclass, field


@dataclass
class XlsxStyleConfig:
    """Injectable style configuration for Excel workbooks."""

    header_bold: bool = True
    header_fill_color: str = "F0F0F0"
    header_font_color: str = "000000"
    header_alignment: str = "center"
    max_column_width: int = 50
    column_padding: int = 3
    freeze_header: bool = True


@dataclass
class Sheet:
    """Typed representation of a worksheet.

    If headers are provided, they are written as the first row with styling.
    Rows follow immediately after headers (or start at row 1 if no headers).
    """

    name: str
    headers: list[str] = field(default_factory=list)
    rows: list[list] = field(default_factory=list)


def parse_sheet(name: str, data: list[dict]) -> Sheet:
    """Convert a list of dicts into a Sheet with headers derived from keys."""
    if not data:
        return Sheet(name=name)
    headers = list(data[0].keys())
    rows = [[d.get(k, "") for k in headers] for d in data]
    return Sheet(name=name, headers=headers, rows=rows)


def generate_xlsx(
    path: str,
    sheets: list[Sheet],
    style: XlsxStyleConfig | None = None,
) -> None:
    """Generate an XLSX workbook from typed Sheet data.

    Args:
        path: Output file path.
        sheets: List of Sheet dataclasses to write.
        style: Optional style config. Uses defaults if None.
    """
    import openpyxl  # lazy import

    cfg = style or XlsxStyleConfig()
    wb = openpyxl.Workbook()

    for i, sheet in enumerate(sheets):
        if i == 0:
            ws = wb.active
            ws.title = sheet.name
        else:
            ws = wb.create_sheet(title=sheet.name)
        _write_sheet(ws, sheet, cfg)

    wb.save(path)


def _write_sheet(ws, sheet: Sheet, cfg: XlsxStyleConfig) -> None:
    """Write a single Sheet to a worksheet."""
    from openpyxl.styles import Alignment, Font, PatternFill  # lazy import

    start_row = 1

    # Write headers with styling
    if sheet.headers:
        header_font = Font(bold=cfg.header_bold, color=cfg.header_font_color)
        header_fill = PatternFill(
            start_color=cfg.header_fill_color,
            end_color=cfg.header_fill_color,
            fill_type="solid",
        )
        header_align = Alignment(horizontal=cfg.header_alignment)

        for j, header in enumerate(sheet.headers):
            cell = ws.cell(row=1, column=j + 1, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        start_row = 2

        if cfg.freeze_header:
            ws.freeze_panes = "A2"

    # Write data rows
    for i, row in enumerate(sheet.rows):
        for j, val in enumerate(row):
            ws.cell(row=start_row + i, column=j + 1, value=val)

    # Auto-width columns based on content
    all_rows = ([sheet.headers] if sheet.headers else []) + sheet.rows
    if not all_rows:
        return
    max_cols = max(len(r) for r in all_rows)
    for col_idx in range(max_cols):
        max_len = 0
        for row in all_rows:
            if col_idx < len(row) and row[col_idx] is not None:
                max_len = max(max_len, len(str(row[col_idx])))
        col_letter = ws.cell(row=1, column=col_idx + 1).column_letter
        ws.column_dimensions[col_letter].width = min(
            max_len + cfg.column_padding, cfg.max_column_width
        )
