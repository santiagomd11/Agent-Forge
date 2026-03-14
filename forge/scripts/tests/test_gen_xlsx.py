"""Tests for XLSX generator."""

import os
import tempfile

import pytest
from openpyxl import load_workbook

from forge.scripts.src.gen_xlsx import (
    generate_xlsx,
    Sheet,
    XlsxStyleConfig,
    parse_sheet,
)


@pytest.fixture
def out_path():
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    yield path
    if os.path.exists(path):
        os.unlink(path)


# --- Sheet dataclass tests ---


class TestSheet:

    def test_defaults(self):
        s = Sheet(name="Data")
        assert s.name == "Data"
        assert s.headers == []
        assert s.rows == []

    def test_with_headers_and_rows(self):
        s = Sheet(name="Users", headers=["Name", "Age"], rows=[["Alice", 30]])
        assert s.headers == ["Name", "Age"]
        assert s.rows == [["Alice", 30]]


# --- parse_sheet tests ---


class TestParseSheet:

    def test_from_dicts(self):
        data = [{"name": "Alice", "score": 95}, {"name": "Bob", "score": 88}]
        s = parse_sheet("Results", data)
        assert s.name == "Results"
        assert s.headers == ["name", "score"]
        assert s.rows == [["Alice", 95], ["Bob", 88]]

    def test_empty_dicts(self):
        s = parse_sheet("Empty", [])
        assert s.headers == []
        assert s.rows == []

    def test_missing_keys(self):
        data = [{"a": 1, "b": 2}, {"a": 3}]
        s = parse_sheet("Sparse", data)
        assert s.headers == ["a", "b"]
        assert s.rows[1] == [3, ""]


# --- XlsxStyleConfig tests ---


class TestXlsxStyleConfig:

    def test_defaults(self):
        cfg = XlsxStyleConfig()
        assert cfg.header_bold is True
        assert cfg.header_fill_color == "F0F0F0"
        assert cfg.max_column_width == 50
        assert cfg.freeze_header is True

    def test_custom_values(self):
        cfg = XlsxStyleConfig(
            header_bold=False,
            header_fill_color="336699",
            max_column_width=30,
            freeze_header=False,
        )
        assert cfg.header_bold is False
        assert cfg.header_fill_color == "336699"
        assert cfg.max_column_width == 30
        assert cfg.freeze_header is False


# --- generate_xlsx tests ---


class TestGenerateXlsx:

    def test_creates_file(self, out_path):
        generate_xlsx(out_path, [Sheet(name="S1", rows=[["A", 1]])])
        assert os.path.isfile(out_path)

    def test_valid_xlsx(self, out_path):
        generate_xlsx(out_path, [Sheet(name="S1", rows=[["A"]])])
        wb = load_workbook(out_path)
        assert wb.active is not None
        wb.close()

    def test_data_correct(self, out_path):
        s = Sheet(name="Data", headers=["Name", "Score"], rows=[["Bob", 95], ["Eve", 88]])
        generate_xlsx(out_path, [s])
        wb = load_workbook(out_path)
        ws = wb.active
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(1, 2).value == "Score"
        assert ws.cell(2, 1).value == "Bob"
        assert ws.cell(2, 2).value == 95
        assert ws.cell(3, 2).value == 88
        wb.close()

    def test_sheet_name(self, out_path):
        generate_xlsx(out_path, [Sheet(name="Results", rows=[["X"]])])
        wb = load_workbook(out_path)
        assert "Results" in wb.sheetnames
        wb.close()

    def test_multiple_sheets(self, out_path):
        sheets = [
            Sheet(name="Users", headers=["Name"], rows=[["Alice"]]),
            Sheet(name="Scores", headers=["Score"], rows=[[100]]),
        ]
        generate_xlsx(out_path, sheets)
        wb = load_workbook(out_path)
        assert "Users" in wb.sheetnames
        assert "Scores" in wb.sheetnames
        assert wb["Users"].cell(2, 1).value == "Alice"
        assert wb["Scores"].cell(2, 1).value == 100
        wb.close()

    def test_empty_rows(self, out_path):
        generate_xlsx(out_path, [Sheet(name="Empty")])
        assert os.path.isfile(out_path)

    def test_bold_header(self, out_path):
        s = Sheet(name="S1", headers=["Name", "Age"], rows=[["Alice", 30]])
        generate_xlsx(out_path, [s])
        wb = load_workbook(out_path)
        ws = wb.active
        assert ws.cell(1, 1).font.bold is True
        wb.close()

    def test_no_bold_header_when_disabled(self, out_path):
        s = Sheet(name="S1", headers=["Name"], rows=[["Alice"]])
        style = XlsxStyleConfig(header_bold=False)
        generate_xlsx(out_path, [s], style=style)
        wb = load_workbook(out_path)
        ws = wb.active
        assert ws.cell(1, 1).font.bold is False
        wb.close()

    def test_from_dicts_via_parse_sheet(self, out_path):
        data = [{"name": "Alice", "score": 95}, {"name": "Bob", "score": 88}]
        s = parse_sheet("Results", data)
        generate_xlsx(out_path, [s])
        wb = load_workbook(out_path)
        ws = wb.active
        assert ws.cell(1, 1).value == "name"
        assert ws.cell(1, 2).value == "score"
        assert ws.cell(2, 1).value == "Alice"
        assert ws.cell(3, 2).value == 88
        wb.close()

    def test_auto_column_width(self, out_path):
        s = Sheet(name="S1", headers=["Short", "A very long column header value"], rows=[["x", "y"]])
        generate_xlsx(out_path, [s])
        wb = load_workbook(out_path)
        ws = wb.active
        assert ws.column_dimensions["B"].width > ws.column_dimensions["A"].width
        wb.close()

    def test_freeze_header(self, out_path):
        s = Sheet(name="S1", headers=["A", "B"], rows=[[1, 2]])
        generate_xlsx(out_path, [s])
        wb = load_workbook(out_path)
        ws = wb.active
        assert ws.freeze_panes == "A2"
        wb.close()

    def test_no_freeze_when_disabled(self, out_path):
        s = Sheet(name="S1", headers=["A"], rows=[[1]])
        style = XlsxStyleConfig(freeze_header=False)
        generate_xlsx(out_path, [s], style=style)
        wb = load_workbook(out_path)
        ws = wb.active
        assert ws.freeze_panes is None
        wb.close()

    def test_custom_fill_color(self, out_path):
        s = Sheet(name="S1", headers=["A"], rows=[[1]])
        style = XlsxStyleConfig(header_fill_color="336699")
        generate_xlsx(out_path, [s], style=style)
        wb = load_workbook(out_path)
        ws = wb.active
        assert ws.cell(1, 1).fill.start_color.rgb == "00336699"
        wb.close()

    def test_rows_only_no_headers(self, out_path):
        s = Sheet(name="Raw", rows=[["A", 1], ["B", 2]])
        generate_xlsx(out_path, [s])
        wb = load_workbook(out_path)
        ws = wb.active
        assert ws.cell(1, 1).value == "A"
        assert ws.cell(1, 1).font.bold is not True
        wb.close()

    def test_custom_max_width(self, out_path):
        s = Sheet(name="S1", headers=["X" * 100], rows=[["y"]])
        style = XlsxStyleConfig(max_column_width=20)
        generate_xlsx(out_path, [s], style=style)
        wb = load_workbook(out_path)
        ws = wb.active
        assert ws.column_dimensions["A"].width <= 20
        wb.close()
